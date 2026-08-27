#!/usr/bin/env python3
"""ATLAS OSFL · loader restringido del padrón oficial RNPJSFL.

Procesa únicamente un archivo local entregado por el workflow autorizado. Soporta
XLSX/XLS/CSV/ZIP/RAR, conserva linaje y usa GitHub OIDC para enviar staging a Supabase.
"""
from __future__ import annotations
import argparse,csv,datetime as dt,hashlib,itertools,json,os,re,tempfile,time,unicodedata,urllib.request,zipfile
from collections import Counter
from pathlib import Path

EDGE="https://ldmtlwzqaqmegedktlxr.supabase.co/functions/v1/aml-osfl-registry-ingest"
AUD="atlas-osfl-registry-ingest";UA="ATLAS-OSFL-RNPJSFL/0.94";TOK={"value":None,"at":0.0}
ALIASES={
"registry_number":{"numero_de_registro","numero_registro","nro_registro","n_registro","numero_de_inscripcion","numero_inscripcion","nro_inscripcion","n_inscripcion","inscripcion","numero"},
"legal_name":{"nombre_persona_juridica","nombre_de_persona_juridica","nombre_pj","razon_social","nombre_organizacion","nombre_de_la_organizacion","nombre"},
"rut":{"rut","rut_persona_juridica","rut_pj","rol_unico_tributario"},"dv":{"dv","digito_verificador"},
"origin":{"origen","organismo_de_origen","organismo_origen","procedencia"},"commune":{"comuna","comuna_domicilio"},"region":{"region","region_domicilio"},
"address":{"direccion","domicilio","direccion_persona_juridica","domicilio_persona_juridica"},
"organization_type":{"tipo_persona_juridica","tipo_de_persona_juridica","tipo_organizacion","tipo_de_organizacion","tipo"},
"classification":{"clasificacion","clasificacion_organizacion","clasificacion_persona_juridica"},
"grant_date":{"fecha_concesion","fecha_de_concesion","fecha_otorgamiento","fecha_personalidad_juridica","fecha_decreto"},
"registration_date":{"fecha_inscripcion","fecha_de_inscripcion","fecha_registro","fecha_de_registro"},
"legal_status":{"estado","estado_vigencia","vigencia","estado_persona_juridica"}}
INACTIVE={"NO_VIGENTE","DISUELTA","DISUELTO","EXTINTA","EXTINTO","CADUCADA","CADUCADO","ANULADA","ANULADO"}
REGIONS=("ARICA","TARAPACA","ANTOFAGASTA","ATACAMA","COQUIMBO","VALPARAISO","METROPOLITANA","OHIGGINS","MAULE","NUBLE","BIOBIO","ARAUCANIA","RIOS","LAGOS","AYSEN","MAGALLANES")

def norm(v):
 x=unicodedata.normalize("NFKD",str(v or "")).encode("ascii","ignore").decode().lower();return re.sub(r"[^a-z0-9]+","_",x).strip("_")
def clean(v):
 if v is None:return None
 if isinstance(v,float) and v.is_integer():v=int(v)
 s=re.sub(r"\s+"," ",str(v)).strip();return s or None
def date_iso(v):
 if v is None or v=="":return None
 if isinstance(v,dt.datetime):return v.date().isoformat()
 if isinstance(v,dt.date):return v.isoformat()
 if isinstance(v,(int,float)) and 15000<=float(v)<=80000:
  try:return (dt.datetime(1899,12,30)+dt.timedelta(days=float(v))).date().isoformat()
  except Exception:return None
 s=clean(v)
 if not s:return None
 for f in ("%Y-%m-%d","%d-%m-%Y","%d/%m/%Y","%d.%m.%Y","%Y/%m/%d"):
  try:return dt.datetime.strptime(s.split(" ")[0],f).date().isoformat()
  except ValueError:pass
 return None
def rut_valid(raw,dv=None):
 s=clean(raw)
 if not s:return None
 a=re.sub(r"[^0-9Kk]","",s).upper()
 if dv is not None and re.sub(r"[^0-9Kk]","",str(dv)):a=re.sub(r"\D","",s)+re.sub(r"[^0-9Kk]","",str(dv)).upper()[:1]
 if len(a)<2:return None
 body,d=a[:-1],a[-1]
 if not body.isdigit() or d not in "0123456789K":return None
 body=body.lstrip("0") or "0";total=0;factor=2
 for ch in reversed(body):total+=int(ch)*factor;factor=2 if factor==7 else factor+1
 r=11-(total%11);expected="0" if r==11 else "K" if r==10 else str(r)
 return f"{int(body)}-{d}" if d==expected else None
def active_status(v):
 n=norm(v).upper()
 if n=="VIGENTE" or n.startswith("VIGENTE_"):return True
 if n in INACTIVE or n.startswith("NO_VIGENTE"):return False
 return None
def sheet_region(name):
 n=norm(name).upper();return clean(name) if any(k in n for k in REGIONS) else None

def token():
 now=time.time()
 if TOK["value"] and now-TOK["at"]<210:return TOK["value"]
 base=os.environ.get("ACTIONS_ID_TOKEN_REQUEST_URL","");secret=os.environ.get("ACTIONS_ID_TOKEN_REQUEST_TOKEN","")
 if not base or not secret:raise RuntimeError("GitHub OIDC unavailable")
 req=urllib.request.Request(base+("&" if "?" in base else "?")+"audience="+AUD,headers={"Authorization":"bearer "+secret,"Accept":"application/json"})
 with urllib.request.urlopen(req,timeout=30) as r:value=json.load(r).get("value")
 if not value:raise RuntimeError("OIDC token missing")
 TOK.update(value=value,at=now);return value
def api(payload,timeout=900):
 raw=json.dumps(payload,ensure_ascii=False,separators=(",",":"),default=str).encode();last=None
 for attempt in range(4):
  try:
   req=urllib.request.Request(EDGE,data=raw,headers={"Authorization":"Bearer "+token(),"Content-Type":"application/json","User-Agent":UA},method="POST")
   with urllib.request.urlopen(req,timeout=timeout) as r:out=json.load(r)
   if not out.get("ok"):raise RuntimeError(out.get("error") or "ingest rejected")
   return out
  except Exception as e:
   last=e
   if attempt==3:raise
   TOK.update(value=None,at=0.0);time.sleep(1.5*(2**attempt))
 raise last

def expand_source(src,tmp):
 ext=src.suffix.lower();out=tmp/"expanded";out.mkdir(exist_ok=True)
 if ext==".zip":
  with zipfile.ZipFile(src) as z:z.extractall(out)
 elif ext==".rar":
  import rarfile
  with rarfile.RarFile(src) as r:r.extractall(out)
 elif ext in {".xlsx",".xls",".csv",".txt"}:return [src]
 else:raise RuntimeError(f"unsupported source format: {ext}")
 files=[p for p in out.rglob("*") if p.is_file() and p.suffix.lower() in {".xlsx",".xls",".csv",".txt"}]
 if not files:raise RuntimeError("archive has no spreadsheet")
 return sorted(files,key=lambda p:p.stat().st_size,reverse=True)
def header_map(values):
 h=[norm(v) for v in values];found={}
 for key,aliases in ALIASES.items():
  for i,x in enumerate(h):
   if x in aliases:found[key]=i;break
 return found if "registry_number" in found and "legal_name" in found and len(found)>=4 else None
def iter_rows(sheet,rows):
 buf=list(itertools.islice(rows,25));idx=mp=None
 for i,row in enumerate(buf):
  m=header_map(row)
  if m:idx=i;mp=m;break
 if mp is None:return
 for rownum,row in enumerate(itertools.chain(buf[idx+1:],rows),start=idx+2):
  if not any(clean(x) for x in row):continue
  def g(k):
   j=mp.get(k);return row[j] if j is not None and j<len(row) else None
  raw=clean(g("rut"));valid=rut_valid(raw,g("dv"));status=clean(g("legal_status"))
  rec={"source_sheet":sheet,"source_row_number":rownum,"registry_number":clean(g("registry_number")),"legal_name":clean(g("legal_name")),"rut_raw":raw,"rut":valid,"rut_is_valid":bool(valid),"origin":clean(g("origin")),"commune":clean(g("commune")),"region":clean(g("region")) or sheet_region(sheet),"address":clean(g("address")),"organization_type":clean(g("organization_type")),"classification":clean(g("classification")),"grant_date":date_iso(g("grant_date")),"registration_date":date_iso(g("registration_date")),"legal_status":status,"is_active":active_status(status)}
  rec["source_record_hash"]=hashlib.sha256(json.dumps(rec,ensure_ascii=False,sort_keys=True,separators=(",",":"),default=str).encode()).hexdigest();yield rec
def workbook(path):
 ext=path.suffix.lower()
 if ext==".xlsx":
  import openpyxl
  wb=openpyxl.load_workbook(path,read_only=True,data_only=True)
  try:
   for ws in wb.worksheets:yield from iter_rows(f"{path.stem}:{ws.title}",(list(r) for r in ws.iter_rows(values_only=True)))
  finally:wb.close()
 elif ext==".xls":
  import xlrd
  book=xlrd.open_workbook(path,on_demand=True)
  try:
   for s in book.sheet_names():
    sh=book.sheet_by_name(s);yield from iter_rows(f"{path.stem}:{s}",(sh.row_values(i) for i in range(sh.nrows)))
  finally:book.release_resources()
 else:
  f=path.open("r",encoding="utf-8-sig",errors="replace",newline="");sample=f.read(20000);f.seek(0)
  try:d=csv.Sniffer().sniff(sample,delimiters=";,\t|")
  except csv.Error:
   class Semi(csv.excel):delimiter=";"
   d=Semi()
  try:yield from iter_rows(path.stem,csv.reader(f,dialect=d))
  finally:f.close()

def ingest(path,snapshot_date,source_url=None,expected=None):
 raw=path.read_bytes();sha=hashlib.sha256(raw).hexdigest();load=f"RNPJSFL_{snapshot_date}_{sha[:16]}"
 api({"action":"begin","run":{"load_id":load,"snapshot_date":snapshot_date,"source_url":source_url,"source_file_name":path.name,"source_sha256":sha,"source_bytes":len(raw),"expected_active_total":expected}})
 c=Counter();seen=Counter();batch=[]
 try:
  with tempfile.TemporaryDirectory() as td:
   files=expand_source(path,Path(td))
   for file in files:
    for r in workbook(file):
     c["observed_rows"]+=1;c["accepted_rows"]+=bool(r.get("registry_number") and r.get("legal_name"));c["active_rows"]+=r.get("is_active") is True;c["inactive_rows"]+=r.get("is_active") is False;c["rows_with_rut"]+=bool(r.get("rut_raw"));c["rows_with_valid_rut"]+=bool(r.get("rut_is_valid"))
     if r.get("registry_number"):seen[r["registry_number"]]+=1
     r["load_id"]=load;batch.append(r)
     if len(batch)>=1200:api({"action":"stage_batch","rows":batch});batch=[]
   if batch:api({"action":"stage_batch","rows":batch})
  c["duplicate_registry_numbers"]=sum(v-1 for v in seen.values() if v>1);ratio=c["accepted_rows"]/max(c["observed_rows"],1)
  if c["observed_rows"]<250000:raise RuntimeError(f"source too small:{c['observed_rows']}")
  if c["active_rows"]<200000:raise RuntimeError(f"active population too small:{c['active_rows']}")
  if ratio<0.98:raise RuntimeError(f"accepted ratio too low:{ratio:.4f}")
  api({"action":"stage_complete","load_id":load,**dict(c),"quality":{"local_accepted_ratio":round(ratio,6),"files":len(files)}})
  result=api({"action":"finalize","load_id":load,"expected_active_total":expected});print(json.dumps({"ok":True,"load_id":load,"source_sha256":sha,"local":dict(c),"finalize":result},ensure_ascii=False));return result
 except Exception as e:
  try:api({"action":"fail","load_id":load,"error":str(e)})
  except Exception:pass
  raise

def self_test():
 import openpyxl
 assert rut_valid("76.086.428-5")=="76086428-5" and rut_valid("76.086.428-0") is None
 with tempfile.TemporaryDirectory() as td:
  p=Path(td)/"sample.xlsx";wb=openpyxl.Workbook();ws=wb.active;ws.title="Metropolitana";ws.append(["titulo"]);ws.append(["N° Inscripción","Nombre Persona Jurídica","RUT","Comuna","Región","Tipo Persona Jurídica","Fecha Inscripción","Estado"]);ws.append(["123","FUNDACION PRUEBA","76.086.428-5","Santiago","Metropolitana","Corporación/ Fundación",dt.date(2020,1,2),"Vigente"]);ws.append(["124","CLUB PRUEBA","12.345.678-9","Santiago","Metropolitana","Comunitaria/ Vecinal","03/02/2021","Disuelta"]);wb.save(p);wb.close();rows=list(workbook(p));assert len(rows)==2 and rows[0]["rut_is_valid"] and rows[0]["is_active"] is True and rows[1]["is_active"] is False
 print("OSFL registry loader self-test OK")
def main():
 ap=argparse.ArgumentParser();ap.add_argument("--source-file");ap.add_argument("--source-url");ap.add_argument("--snapshot-date");ap.add_argument("--expected-active-total",type=int);ap.add_argument("--self-test",action="store_true");args=ap.parse_args()
 if args.self_test:return self_test()
 if not args.source_file or not args.snapshot_date:raise SystemExit("--source-file and --snapshot-date are required")
 dt.date.fromisoformat(args.snapshot_date);p=Path(args.source_file)
 if not p.is_file():raise SystemExit(f"source file not found: {p}")
 return ingest(p,args.snapshot_date,args.source_url,args.expected_active_total)
if __name__=="__main__":main()
